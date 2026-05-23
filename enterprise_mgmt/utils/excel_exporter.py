# -*- coding: utf-8 -*-
"""
Excel 导出工具 (utils/excel_exporter.py)
使用 openpyxl 将任意列表数据导出为带样式的 .xlsx 文件。
"""
from __future__ import annotations
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def export_to_excel(
    filepath: str,
    headers: list[str],
    rows: list[list],
    sheet_name: str = "Sheet1",
    title: str = "",
) -> None:
    """
    将数据导出为 Excel 文件。

    :param filepath:   输出文件路径（含 .xlsx 后缀）
    :param headers:    列标题列表
    :param rows:       二维数据列表
    :param sheet_name: 工作表名称
    :param title:      可选的表头大标题（占第一行）
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("请先安装 openpyxl：pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # ── 颜色定义 ─────────────────────────────────────────────────
    HEADER_FILL  = PatternFill("solid", fgColor="2E4057")   # 深蓝灰
    ALT_ROW_FILL = PatternFill("solid", fgColor="F0F4F8")   # 浅蓝灰交替行
    HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
    TITLE_FONT   = Font(bold=True, size=13, color="2E4057")
    THIN_BORDER  = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT   = Alignment(horizontal="left",   vertical="center")

    row_offset = 0

    # ── 可选大标题 ───────────────────────────────────────────────
    if title:
        ws.merge_cells(
            start_row=1, start_column=1,
            end_row=1,   end_column=len(headers)
        )
        cell = ws.cell(row=1, column=1, value=title)
        cell.font      = TITLE_FONT
        cell.alignment = CENTER
        ws.row_dimensions[1].height = 28
        row_offset = 1

    # ── 表头行 ───────────────────────────────────────────────────
    header_row = 1 + row_offset
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
    ws.row_dimensions[header_row].height = 22

    # ── 数据行 ───────────────────────────────────────────────────
    for row_idx, row_data in enumerate(rows, start=header_row + 1):
        fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = LEFT
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill

    # ── 自动列宽 ─────────────────────────────────────────────────
    for col_idx in range(1, len(headers) + 1):
        col_letter  = get_column_letter(col_idx)
        max_len     = len(str(headers[col_idx - 1]))
        for row_data in rows:
            try:
                cell_len = len(str(row_data[col_idx - 1]))
                max_len  = max(max_len, cell_len)
            except IndexError:
                pass
        # 中文字符占约 2 个英文字符宽度，做简单估算
        ws.column_dimensions[col_letter].width = min(max_len * 1.6 + 4, 40)

    # ── 冻结首行（含标题则冻结第二行）───────────────────────────
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    wb.save(filepath)
    logger.info("Excel 导出完成: %s  (%d 行数据)", filepath, len(rows))
