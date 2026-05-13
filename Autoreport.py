“””
Python 自动化报表生成系统
功能：模拟销售数据 → 多维分析 → 自动生成带图表、格式化的 Excel 报表
类比 VBA 的完整自动化流程
“””

import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import pandas as pd

# ─────────────────────────────────────────────

# 1. 生成模拟销售原始数据

# ─────────────────────────────────────────────

def generate_raw_data(n=200):
random.seed(42)
regions   = [“华北”, “华东”, “华南”, “西南”, “西北”]
products  = [“产品A”, “产品B”, “产品C”, “产品D”]
channels  = [“线上”, “线下”, “分销”]
salesreps = [“张伟”, “李娜”, “王芳”, “赵磊”, “陈静”, “刘洋”, “周敏”, “吴强”]

```
start = date(2024, 1, 1)
rows = []
for i in range(1, n + 1):
    qty   = random.randint(10, 200)
    price = random.choice([299, 499, 799, 1299])
    cost  = round(price * random.uniform(0.4, 0.65), 2)
    rows.append({
        "订单ID":   f"ORD-{i:04d}",
        "日期":     start + timedelta(days=random.randint(0, 364)),
        "地区":     random.choice(regions),
        "产品":     random.choice(products),
        "渠道":     random.choice(channels),
        "销售员":   random.choice(salesreps),
        "数量":     qty,
        "单价":     price,
        "销售额":   qty * price,
        "成本":     round(qty * cost, 2),
        "利润":     round(qty * (price - cost), 2),
    })
df = pd.DataFrame(rows)
df["利润率"] = (df["利润"] / df["销售额"] * 100).round(2)
df["季度"]   = df["日期"].apply(lambda d: f"Q{(d.month - 1) // 3 + 1}")
df["月份"]   = df["日期"].apply(lambda d: d.month)
return df
```

# ─────────────────────────────────────────────

# 2. 样式工具函数

# ─────────────────────────────────────────────

BLUE_DARK   = “1F3864”
BLUE_MID    = “2E75B6”
BLUE_LIGHT  = “BDD7EE”
BLUE_PALE   = “DEEAF1”
ORANGE      = “ED7D31”
WHITE       = “FFFFFF”
GRAY_LIGHT  = “F2F2F2”
GRAY_MID    = “D9D9D9”
GREEN       = “70AD47”
RED         = “FF0000”

def thin_border(top=True, bottom=True, left=True, right=True):
s = Side(style=“thin”, color=“BFBFBF”)
n = Side(style=None)
return Border(
top=s if top else n,
bottom=s if bottom else n,
left=s if left else n,
right=s if right else n,
)

def header_style(ws, row, col, value, bg=BLUE_DARK, fg=WHITE,
bold=True, size=11, h_align=“center”):
c = ws.cell(row=row, column=col, value=value)
c.font      = Font(name=“Arial”, bold=bold, color=fg, size=size)
c.fill      = PatternFill(“solid”, fgColor=bg)
c.alignment = Alignment(horizontal=h_align, vertical=“center”, wrap_text=True)
c.border    = thin_border()
return c

def data_style(ws, row, col, value, bg=WHITE, bold=False,
h_align=“center”, num_format=None):
c = ws.cell(row=row, column=col, value=value)
c.font      = Font(name=“Arial”, bold=bold, color=“000000”, size=10)
c.fill      = PatternFill(“solid”, fgColor=bg)
c.alignment = Alignment(horizontal=h_align, vertical=“center”)
c.border    = thin_border()
if num_format:
c.number_format = num_format
return c

def set_col_width(ws, widths: dict):
for col_letter, w in widths.items():
ws.column_dimensions[col_letter].width = w

def freeze_and_filter(ws, freeze_cell, filter_range):
ws.freeze_panes = freeze_cell
ws.auto_filter.ref = filter_range

# ─────────────────────────────────────────────

# 3. Sheet 1 — 封面

# ─────────────────────────────────────────────

def build_cover(wb, report_date):
ws = wb.active
ws.title = “封面”
ws.sheet_view.showGridLines = False

```
# 深色背景
for r in range(1, 40):
    for c in range(1, 15):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=BLUE_DARK)

# 标题
ws.merge_cells("B4:M6")
c = ws["B4"]
c.value     = "年度销售自动化报表"
c.font      = Font(name="Arial", bold=True, size=36, color=WHITE)
c.alignment = Alignment(horizontal="center", vertical="center")

# 副标题
ws.merge_cells("B7:M8")
c = ws["B7"]
c.value     = "Annual Sales Automated Report"
c.font      = Font(name="Arial", size=16, color=BLUE_LIGHT)
c.alignment = Alignment(horizontal="center", vertical="center")

# 橙色分割线
for col in range(2, 14):
    ws.cell(row=9, column=col).fill = PatternFill("solid", fgColor=ORANGE)
ws.row_dimensions[9].height = 4

# 报告信息
info = [
    ("报告期间", "2024年度（1月 - 12月）"),
    ("报告日期", report_date.strftime("%Y年%m月%d日")),
    ("数据来源", "销售管理系统"),
    ("编制单位", "销售分析部"),
]
for i, (label, val) in enumerate(info, start=11):
    ws.merge_cells(f"D{i}:E{i}")
    ws.merge_cells(f"F{i}:K{i}")
    lc = ws[f"D{i}"]
    lc.value     = label
    lc.font      = Font(name="Arial", bold=True, color=ORANGE, size=12)
    lc.alignment = Alignment(horizontal="right", vertical="center")
    vc = ws[f"F{i}"]
    vc.value     = val
    vc.font      = Font(name="Arial", color=WHITE, size=12)
    vc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[i].height = 22

# 目录
ws.merge_cells("B17:M17")
c = ws["B17"]
c.value     = "报 告 目 录"
c.font      = Font(name="Arial", bold=True, size=14, color=ORANGE)
c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[17].height = 28

toc = [
    ("Sheet 1", "封面"),
    ("Sheet 2", "原始数据"),
    ("Sheet 3", "总体概览（KPI + 图表）"),
    ("Sheet 4", "地区分析"),
    ("Sheet 5", "产品分析"),
    ("Sheet 6", "月度趋势"),
    ("Sheet 7", "销售员排名"),
]
for i, (sheet, desc) in enumerate(toc, start=19):
    ws.merge_cells(f"D{i}:E{i}")
    ws.merge_cells(f"F{i}:K{i}")
    ws[f"D{i}"].value = sheet
    ws[f"D{i}"].font  = Font(name="Arial", color=ORANGE, size=11, bold=True)
    ws[f"D{i}"].alignment = Alignment(horizontal="right")
    ws[f"F{i}"].value = desc
    ws[f"F{i}"].font  = Font(name="Arial", color=WHITE, size=11)
    ws[f"F{i}"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[i].height = 20

# 行高
for r in [1, 2, 3]: ws.row_dimensions[r].height = 15
ws.row_dimensions[4].height = 60
ws.row_dimensions[7].height = 30
set_col_width(ws, {get_column_letter(c): 10 for c in range(1, 15)})
ws.column_dimensions["B"].width = 3
ws.column_dimensions["M"].width = 3
```

# ─────────────────────────────────────────────

# 4. Sheet 2 — 原始数据

# ─────────────────────────────────────────────

def build_raw_data(wb, df):
ws = wb.create_sheet(“原始数据”)
ws.sheet_view.showGridLines = False

```
cols = ["订单ID", "日期", "地区", "产品", "渠道", "销售员",
        "数量", "单价", "销售额", "成本", "利润", "利润率", "季度", "月份"]

# 列标题
for j, col in enumerate(cols, 1):
    header_style(ws, 1, j, col, bg=BLUE_MID)

# 数据行
money_fmt   = '#,##0'
percent_fmt = '0.00"%"'
for i, row in enumerate(df[cols].itertuples(index=False), 2):
    bg = WHITE if i % 2 == 0 else GRAY_LIGHT
    for j, val in enumerate(row, 1):
        col_name = cols[j - 1]
        if col_name in ("销售额", "成本", "利润", "单价"):
            data_style(ws, i, j, val, bg=bg, num_format=money_fmt)
        elif col_name == "利润率":
            data_style(ws, i, j, val, bg=bg, num_format=percent_fmt)
        elif col_name == "日期":
            data_style(ws, i, j, val, bg=bg, num_format="yyyy-mm-dd")
        else:
            data_style(ws, i, j, val, bg=bg)

# 冻结 + 筛选
freeze_and_filter(ws, "A2", f"A1:{get_column_letter(len(cols))}1")
ws.row_dimensions[1].height = 30

widths = {"A":12,"B":12,"C":8,"D":8,"E":8,"F":8,
          "G":8,"H":8,"I":12,"J":12,"K":12,"L":10,"M":6,"N":6}
set_col_width(ws, widths)

# 条件格式 — 利润率数据条（L列）
last_row = len(df) + 1
ws.conditional_formatting.add(
    f"L2:L{last_row}",
    DataBarRule(start_type="min", end_type="max",
                color="2E75B6", showValue=True)
)
```

# ─────────────────────────────────────────────

# 5. Sheet 3 — 总体概览

# ─────────────────────────────────────────────

def build_overview(wb, df):
ws = wb.create_sheet(“总体概览”)
ws.sheet_view.showGridLines = False

```
# ── KPI 卡片 ──
kpis = [
    ("总销售额", f"¥{df['销售额'].sum():,.0f}", BLUE_MID),
    ("总利润",   f"¥{df['利润'].sum():,.0f}",   GREEN),
    ("平均利润率", f"{df['利润率'].mean():.1f}%", ORANGE),
    ("订单总数", f"{len(df):,} 单",              BLUE_DARK),
]

ws.merge_cells("A1:Q1")
t = ws["A1"]
t.value     = "📊  2024 年度销售总体概览"
t.font      = Font(name="Arial", bold=True, size=16, color=BLUE_DARK)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 35

card_cols = [1, 5, 9, 13]
for idx, ((title, val, color), start_col) in enumerate(zip(kpis, card_cols)):
    end_col = start_col + 3
    # 标题行
    ws.merge_cells(start_row=3, start_column=start_col,
                   end_row=3, end_column=end_col)
    c = ws.cell(row=3, column=start_col, value=title)
    c.font      = Font(name="Arial", bold=True, color=WHITE, size=11)
    c.fill      = PatternFill("solid", fgColor=color)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # 数值行
    ws.merge_cells(start_row=4, start_column=start_col,
                   end_row=5, end_column=end_col)
    v = ws.cell(row=4, column=start_col, value=val)
    v.font      = Font(name="Arial", bold=True, color=color, size=18)
    v.fill      = PatternFill("solid", fgColor=GRAY_LIGHT)
    v.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[4].height = 30
    ws.row_dimensions[5].height = 30

# ── 季度汇总表 ──
q_data = df.groupby("季度").agg(
    销售额=("销售额", "sum"),
    利润=("利润", "sum"),
    订单数=("订单ID", "count")
).reset_index().sort_values("季度")

r0 = 7
header_style(ws, r0, 1, "季度销售汇总", bg=BLUE_MID, size=12)
ws.merge_cells(f"A{r0}:D{r0}")

headers = ["季度", "销售额(¥)", "利润(¥)", "订单数"]
for j, h in enumerate(headers, 1):
    header_style(ws, r0+1, j, h, bg=BLUE_DARK)

for i, row in enumerate(q_data.itertuples(index=False), r0+2):
    bg = BLUE_PALE if i % 2 == 0 else WHITE
    data_style(ws, i, 1, row.季度, bg=bg)
    data_style(ws, i, 2, row.销售额, bg=bg, num_format='#,##0')
    data_style(ws, i, 3, row.利润,   bg=bg, num_format='#,##0')
    data_style(ws, i, 4, row.订单数, bg=bg)

# 合计行
total_r = r0 + 2 + len(q_data)
data_style(ws, total_r, 1, "合计", bg=BLUE_LIGHT, bold=True)
data_style(ws, total_r, 2, f"=SUM(B{r0+2}:B{total_r-1})",
           bg=BLUE_LIGHT, bold=True, num_format='#,##0')
data_style(ws, total_r, 3, f"=SUM(C{r0+2}:C{total_r-1})",
           bg=BLUE_LIGHT, bold=True, num_format='#,##0')
data_style(ws, total_r, 4, f"=SUM(D{r0+2}:D{total_r-1})",
           bg=BLUE_LIGHT, bold=True)

# ── 柱状图：季度销售额 vs 利润 ──
chart = BarChart()
chart.type    = "col"
chart.title   = "季度销售额 vs 利润"
chart.y_axis.title = "金额 (¥)"
chart.x_axis.title = "季度"
chart.style   = 10
chart.width   = 16
chart.height  = 10

data_ref = Reference(ws, min_col=2, max_col=3,
                     min_row=r0+1, max_row=r0+1+len(q_data))
cats_ref = Reference(ws, min_col=1,
                     min_row=r0+2, max_row=r0+1+len(q_data))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = BLUE_MID
chart.series[1].graphicalProperties.solidFill = GREEN
ws.add_chart(chart, "F7")

# 列宽
for col in range(1, 18):
    ws.column_dimensions[get_column_letter(col)].width = 13
```

# ─────────────────────────────────────────────

# 6. Sheet 4 — 地区分析

# ─────────────────────────────────────────────

def build_region(wb, df):
ws = wb.create_sheet(“地区分析”)
ws.sheet_view.showGridLines = False

```
region_df = df.groupby("地区").agg(
    销售额=("销售额", "sum"),
    利润=("利润", "sum"),
    订单数=("订单ID", "count"),
    平均利润率=("利润率", "mean")
).reset_index().sort_values("销售额", ascending=False)
region_df["利润率pct"] = region_df["平均利润率"].round(1)

ws.merge_cells("A1:F1")
t = ws["A1"]
t.value     = "地区销售分析"
t.font      = Font(name="Arial", bold=True, size=14, color=BLUE_DARK)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

headers = ["地区", "销售额(¥)", "利润(¥)", "订单数", "平均利润率(%)"]
for j, h in enumerate(headers, 1):
    header_style(ws, 2, j, h, bg=BLUE_MID)

for i, row in enumerate(region_df.itertuples(index=False), 3):
    bg = WHITE if i % 2 == 0 else GRAY_LIGHT
    data_style(ws, i, 1, row.地区,    bg=bg, bold=True)
    data_style(ws, i, 2, row.销售额,  bg=bg, num_format='#,##0')
    data_style(ws, i, 3, row.利润,    bg=bg, num_format='#,##0')
    data_style(ws, i, 4, row.订单数,  bg=bg)
    data_style(ws, i, 5, row.利润率pct, bg=bg, num_format='0.0')

last_r = 2 + len(region_df)

# 颜色热力图（销售额列）
ws.conditional_formatting.add(
    f"B3:B{last_r}",
    ColorScaleRule(
        start_type="min", start_color="DEEAF1",
        end_type="max",   end_color="1F3864"
    )
)

# 饼图：销售额占比
pie = PieChart()
pie.title  = "各地区销售额占比"
pie.style  = 10
pie.width  = 14
pie.height = 10

data_ref = Reference(ws, min_col=2, min_row=2, max_row=last_r)
cats_ref = Reference(ws, min_col=1, min_row=3, max_row=last_r)
pie.add_data(data_ref, titles_from_data=True)
pie.set_categories(cats_ref)
ws.add_chart(pie, "G2")

set_col_width(ws, {"A":10,"B":14,"C":14,"D":10,"E":14})
```

# ─────────────────────────────────────────────

# 7. Sheet 5 — 产品分析

# ─────────────────────────────────────────────

def build_product(wb, df):
ws = wb.create_sheet(“产品分析”)
ws.sheet_view.showGridLines = False

```
prod_df = df.groupby("产品").agg(
    销售额=("销售额", "sum"),
    利润=("利润", "sum"),
    数量=("数量", "sum"),
    订单数=("订单ID", "count")
).reset_index().sort_values("销售额", ascending=False)
prod_df["利润率pct"] = (prod_df["利润"] / prod_df["销售额"] * 100).round(1)

# 渠道 × 产品 交叉表
cross = df.pivot_table(index="渠道", columns="产品",
                       values="销售额", aggfunc="sum").fillna(0)

ws.merge_cells("A1:G1")
t = ws["A1"]
t.value     = "产品销售分析"
t.font      = Font(name="Arial", bold=True, size=14, color=BLUE_DARK)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

headers = ["产品", "销售额(¥)", "利润(¥)", "销量", "订单数", "利润率(%)"]
for j, h in enumerate(headers, 1):
    header_style(ws, 2, j, h, bg=ORANGE)

for i, row in enumerate(prod_df.itertuples(index=False), 3):
    bg = WHITE if i % 2 == 0 else GRAY_LIGHT
    data_style(ws, i, 1, row.产品,    bg=bg, bold=True)
    data_style(ws, i, 2, row.销售额,  bg=bg, num_format='#,##0')
    data_style(ws, i, 3, row.利润,    bg=bg, num_format='#,##0')
    data_style(ws, i, 4, row.数量,    bg=bg)
    data_style(ws, i, 5, row.订单数,  bg=bg)
    data_style(ws, i, 6, row.利润率pct, bg=bg, num_format='0.0"%"')

# 柱状图：产品销售额
chart = BarChart()
chart.type    = "col"
chart.title   = "产品销售额对比"
chart.style   = 10
chart.width   = 14
chart.height  = 10
last_r        = 2 + len(prod_df)
data_ref = Reference(ws, min_col=2, min_row=2, max_row=last_r)
cats_ref = Reference(ws, min_col=1, min_row=3, max_row=last_r)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = ORANGE
ws.add_chart(chart, "H2")

# 渠道交叉表
r0 = last_r + 2
ws.merge_cells(f"A{r0}:E{r0}")
header_style(ws, r0, 1, "渠道 × 产品 销售额交叉分析", bg=BLUE_MID)

# 表头
header_style(ws, r0+1, 1, "渠道", bg=BLUE_DARK)
for j, col_name in enumerate(cross.columns, 2):
    header_style(ws, r0+1, j, col_name, bg=BLUE_DARK)

for i, (ch, row_data) in enumerate(cross.iterrows(), r0+2):
    bg = WHITE if i % 2 == 0 else GRAY_LIGHT
    data_style(ws, i, 1, ch, bg=bg, bold=True)
    for j, val in enumerate(row_data, 2):
        data_style(ws, i, j, val, bg=bg, num_format='#,##0')

set_col_width(ws, {get_column_letter(c): 13 for c in range(1, 10)})
```

# ─────────────────────────────────────────────

# 8. Sheet 6 — 月度趋势

# ─────────────────────────────────────────────

def build_monthly(wb, df):
ws = wb.create_sheet(“月度趋势”)
ws.sheet_view.showGridLines = False

```
monthly = df.groupby("月份").agg(
    销售额=("销售额", "sum"),
    利润=("利润", "sum"),
    订单数=("订单ID", "count")
).reset_index().sort_values("月份")
monthly["月份标签"] = monthly["月份"].apply(lambda m: f"{m}月")

ws.merge_cells("A1:E1")
t = ws["A1"]
t.value     = "月度销售趋势分析"
t.font      = Font(name="Arial", bold=True, size=14, color=BLUE_DARK)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

headers = ["月份", "销售额(¥)", "利润(¥)", "订单数", "环比增长"]
for j, h in enumerate(headers, 1):
    header_style(ws, 2, j, h, bg=BLUE_MID)

for i, row in enumerate(monthly.itertuples(index=False), 3):
    bg = WHITE if i % 2 == 0 else GRAY_LIGHT
    data_style(ws, i, 1, row.月份标签, bg=bg, bold=True)
    data_style(ws, i, 2, row.销售额,   bg=bg, num_format='#,##0')
    data_style(ws, i, 3, row.利润,     bg=bg, num_format='#,##0')
    data_style(ws, i, 4, row.订单数,   bg=bg)

# 环比增长公式（从第4行起有上一行可比）
for i in range(4, 3 + len(monthly)):
    c = ws.cell(row=i, column=5,
                value=f"=(B{i}-B{i-1})/B{i-1}")
    c.number_format = "0.0%"
    c.font          = Font(name="Arial", size=10)
    c.alignment     = Alignment(horizontal="center")
    c.border        = thin_border()
ws.cell(row=3, column=5, value="—")

last_r = 2 + len(monthly)

# 折线图：月度趋势
line = LineChart()
line.title   = "月度销售额 & 利润趋势"
line.style   = 10
line.y_axis.title = "金额 (¥)"
line.x_axis.title = "月份"
line.width   = 22
line.height  = 12

data_ref = Reference(ws, min_col=2, max_col=3,
                     min_row=2, max_row=last_r)
cats_ref = Reference(ws, min_col=1, min_row=3, max_row=last_r)
line.add_data(data_ref, titles_from_data=True)
line.set_categories(cats_ref)
line.series[0].graphicalProperties.line.solidFill = BLUE_MID
line.series[0].graphicalProperties.line.width     = 20000
line.series[1].graphicalProperties.line.solidFill = GREEN
line.series[1].graphicalProperties.line.width     = 20000
ws.add_chart(line, "G2")

set_col_width(ws, {"A":8,"B":14,"C":14,"D":10,"E":12})
for col in range(7, 20):
    ws.column_dimensions[get_column_letter(col)].width = 10
```

# ─────────────────────────────────────────────

# 9. Sheet 7 — 销售员排名

# ─────────────────────────────────────────────

def build_ranking(wb, df):
ws = wb.create_sheet(“销售员排名”)
ws.sheet_view.showGridLines = False

```
rep_df = df.groupby("销售员").agg(
    销售额=("销售额", "sum"),
    利润=("利润", "sum"),
    订单数=("订单ID", "count"),
    平均利润率=("利润率", "mean")
).reset_index().sort_values("销售额", ascending=False).reset_index(drop=True)
rep_df.index += 1  # 排名从 1 开始

ws.merge_cells("A1:F1")
t = ws["A1"]
t.value     = "销售员业绩排名"
t.font      = Font(name="Arial", bold=True, size=14, color=BLUE_DARK)
t.alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[1].height = 30

headers = ["排名", "销售员", "销售额(¥)", "利润(¥)", "订单数", "平均利润率(%)"]
for j, h in enumerate(headers, 1):
    header_style(ws, 2, j, h, bg=BLUE_DARK)

medal = {1: "🥇", 2: "🥈", 3: "🥉"}
for rank, row in rep_df.iterrows():
    bg = {1: "FFD700", 2: "C0C0C0", 3: "CD7F32"}.get(rank, WHITE if rank % 2 == 0 else GRAY_LIGHT)
    bold = rank <= 3
    data_style(ws, rank+2, 1, f"{medal.get(rank,'')} {rank}", bg=bg, bold=bold)
    data_style(ws, rank+2, 2, row["销售员"],   bg=bg, bold=bold)
    data_style(ws, rank+2, 3, row["销售额"],   bg=bg, bold=bold, num_format='#,##0')
    data_style(ws, rank+2, 4, row["利润"],     bg=bg, bold=bold, num_format='#,##0')
    data_style(ws, rank+2, 5, row["订单数"],   bg=bg, bold=bold)
    data_style(ws, rank+2, 6, round(row["平均利润率"], 1), bg=bg, bold=bold, num_format='0.0')

last_r = 2 + len(rep_df)

# 横向条形图：销售员业绩
chart = BarChart()
chart.type    = "bar"
chart.title   = "销售员销售额排名"
chart.style   = 10
chart.width   = 16
chart.height  = 14
data_ref = Reference(ws, min_col=3, min_row=2, max_row=last_r)
cats_ref = Reference(ws, min_col=2, min_row=3, max_row=last_r)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = BLUE_MID
ws.add_chart(chart, "H2")

set_col_width(ws, {"A":10,"B":10,"C":14,"D":14,"E":10,"F":14})
```

# ─────────────────────────────────────────────

# 10. 统一美化：标签颜色 + 页眉页脚

# ─────────────────────────────────────────────

TAB_COLORS = {
“封面”:     “1F3864”,
“原始数据”: “595959”,
“总体概览”: “2E75B6”,
“地区分析”: “70AD47”,
“产品分析”: “ED7D31”,
“月度趋势”: “4472C4”,
“销售员排名”:“FFC000”,
}

def finalize_workbook(wb):
for ws in wb.worksheets:
# Sheet 标签颜色
color = TAB_COLORS.get(ws.title, “000000”)
ws.sheet_properties.tabColor = color

```
    # 页眉页脚
    ws.oddHeader.center.text = f"&\"Arial,Bold\"&14 2024年度销售报表 — {ws.title}"
    ws.oddHeader.center.font = "Arial"
    ws.oddFooter.left.text   = "&\"Arial\"&9 机密文件，请勿外传"
    ws.oddFooter.right.text  = "&\"Arial\"&9 第 &P 页 / 共 &N 页"

    # 打印设置
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.page_setup.fitToHeight = 0
```

# ─────────────────────────────────────────────

# 11. 自检函数

# ─────────────────────────────────────────────

def self_check(wb, df):
issues = []

```
expected_sheets = ["封面", "原始数据", "总体概览", "地区分析",
                    "产品分析", "月度趋势", "销售员排名"]
for s in expected_sheets:
    if s not in wb.sheetnames:
        issues.append(f"❌ 缺少 Sheet: {s}")

raw_ws = wb["原始数据"]
if raw_ws.max_row - 1 != len(df):
    issues.append(f"❌ 原始数据行数不匹配: 期望{len(df)}, 实际{raw_ws.max_row-1}")

for sheet_name in expected_sheets:
    ws = wb[sheet_name]
    # 封面用merge_cells，内容在B4；其他在A1
    check_cell = "B4" if sheet_name == "封面" else "A1"
    if ws[check_cell].value is None:
        issues.append(f"⚠️ Sheet '{sheet_name}' {check_cell} 为空")

# 检查总体概览 KPI 不为空（KPI写在merge首格，用B4 col=1即A4位置，实际存A4/B4）
overview = wb["总体概览"]
# KPI 卡片标题在 row=3, col=1
if overview.cell(row=3, column=1).value is None:
    issues.append("❌ 总体概览 KPI 卡片为空")

if issues:
    print("\n🔍 自检发现问题：")
    for iss in issues:
        print(f"  {iss}")
    return False
else:
    print("\n✅ 自检通过：所有 Sheet 和数据完整，报表结构正常。")
    return True
```

# ─────────────────────────────────────────────

# 12. 主程序入口

# ─────────────────────────────────────────────

def main():
output_path = “/mnt/user-data/outputs/销售自动化报表_2024.xlsx”
today       = date.today()

```
print("=" * 55)
print("  Python 自动化报表系统  |  类比 VBA 完整功能")
print("=" * 55)

print("▶ [1/8] 生成模拟销售数据...")
df = generate_raw_data(200)
print(f"     共生成 {len(df)} 条销售记录，涵盖 {df['地区'].nunique()} 个地区、"
      f"{df['产品'].nunique()} 款产品")

print("▶ [2/8] 创建工作簿 & 封面...")
wb = Workbook()
build_cover(wb, today)

print("▶ [3/8] 写入原始数据（含条件格式）...")
build_raw_data(wb, df)

print("▶ [4/8] 生成总体概览（KPI + 季度柱状图）...")
build_overview(wb, df)

print("▶ [5/8] 生成地区分析（热力图 + 饼图）...")
build_region(wb, df)

print("▶ [6/8] 生成产品分析（对比图 + 交叉表）...")
build_product(wb, df)

print("▶ [7/8] 生成月度趋势（折线图 + 环比公式）...")
build_monthly(wb, df)

print("▶ [8/8] 生成销售员排名（奖牌 + 条形图）...")
build_ranking(wb, df)

print("\n▶ 统一美化（标签颜色 + 页眉页脚 + 打印设置）...")
finalize_workbook(wb)

print("▶ 执行自检...")
ok = self_check(wb, df)

print(f"\n▶ 保存报表到：{output_path}")
wb.save(output_path)

print("\n" + "=" * 55)
if ok:
    print("  🎉 报表生成成功！")
else:
    print("  ⚠️ 报表已生成，但存在上述问题，请检查。")
print(f"  📁 文件路径：{output_path}")
print("=" * 55)

# 输出关键统计
print(f"\n📊 数据摘要：")
print(f"   总销售额：¥{df['销售额'].sum():>15,.0f}")
print(f"   总利润  ：¥{df['利润'].sum():>15,.0f}")
print(f"   平均利润率：{df['利润率'].mean():>10.1f}%")
print(f"   最佳地区：{df.groupby('地区')['销售额'].sum().idxmax()}")
print(f"   最佳产品：{df.groupby('产品')['销售额'].sum().idxmax()}")
top_rep = df.groupby('销售员')['销售额'].sum().idxmax()
print(f"   销售冠军：{top_rep}")
```

if **name** == “**main**”:
main()
